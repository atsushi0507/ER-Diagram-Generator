from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

def make_sheet(wb):
    ws = wb.active
    ws.title="table_1"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 35

    ws.cell(1, 1, value="Column Name")
    ws.cell(1, 2, value="Primary Key")
    ws.cell(1, 3, value="Foreign Key")
    ws.cell(1, 4, value="Comment")
    ws["F1"].value = "URL"

    return ws

def make_table(title):
    table = Table(displayName=title, ref=f"A1:D2")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    return table

def make_empty_book(filename="sample_format.xlsx", n_tables=4):
    wb = Workbook()
    ws = make_sheet(wb)
    ws.add_table(make_table("table_1"))
    for n in range(2, n_tables+1):
        ws = wb.copy_worksheet(wb["table_1"])
        ws.title = f"table_{n}"
        table = make_table(f"table_{n}")
        ws.add_table(table)

    ws_relation = wb.create_sheet(title="relation")
    ws_relation.column_dimensions["A"].width = 30
    ws_relation.column_dimensions["B"].width = 30
    ws_relation.column_dimensions["C"].width = 40

    ws_relation.cell(1, 1, value="From")
    ws_relation.cell(1, 2, value="To")
    ws_relation.cell(1, 3, value="Description")

    table2 = Table(displayName="Relation", ref="A1:C6")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table2.tableStyleInfo = style

    ws_relation.add_table(table2)

    """
    ws_process = wb.create_sheet(title="process")
    ws_process.column_dimensions["A"].width = 12
    ws_process.column_dimensions["B"].width = 25
    ws_process.column_dimensions["C"].width = 25
    ws_process.column_dimensions["D"].width = 35

    ws_process.cell(1, 1, value="Process")
    ws_process.cell(1, 2, value="Input")
    ws_process.cell(1, 3, value="Output")
    ws_process.cell(1, 4, value="Detail")

    table3 = Table(displayName="Process", ref="A1:D2")
    table3.tableStyleInfo = style
    ws_process.add_table(table3)
    
    dv = DataValidation(
        type="list",
        formula1='"join,union,process,middle_table"',
        allow_blank=True
    )
    dv.add(f"A2:A1000")
    ws_process.add_data_validation(dv)
    """
    
    wb.save(filename)
    wb.close()
    print(f"Generated empty excel book: {filename}")


def generate_excelbook(filename="sample_format.xlsx"):
    wb = Workbook()
    ws = make_sheet(wb)
    ws.add_table(make_table("table_1"))
    dummy_data = {
        "table_1": {
            "columns": [f"column_{i}" for i in range(10)],
            "PKs": [1, 0, 1, 0, 0, 1, 0, 0, 0, 0]
        },
        "table_2": {
            "columns": [f"column_{i}" for i in range(7)],
            "PKs": [0, 1, 0, 0, 0, 0, 0]
        },
        "table_3": {
            "columns": [f"column_{i}" for i in range(4)],
            "PKs": [1, 1, 0, 0]
        },
        "table_4": {
            "columns": [f"column_{i}" for i in range(6)],
            "PKs": [1, 0, 0, 0, 0, 0]
        },
        "table_5": {
            "columns": [f"column_{i}" for i in range(5)],
            "PKs": [1, 1, 0, 1, 0]
        },
        "table_6": {
            "columns": [f"column_{i}" for i in range(4)],
            "PKs": [1, 1, 0, 0]
        },
        "table_7": {
            "columns": [f"column_{i}" for i in range(3)],
            "PKs": [1, 0, 0]
        }
    }
    for table, values in dummy_data.items():
        if not table == "table_1":
            ws = wb.copy_worksheet(wb["table_1"])
            ws.title = f"{table}"
            for key in values.keys():
                for i in range(len(dummy_data[table][key])):
                    if key == "columns":
                        ws.cell(i+2, 1, value=dummy_data[table][key][i])
                    elif key == "PKs":
                        ws.cell(i+2, 2, value=dummy_data[table][key][i])
            ws["F1"].value = "URL"
            table = make_table(f"{table}")
            ws.add_table(table)
    ws = wb["table_1"]
    for i in range(len(dummy_data["table_1"]["columns"])):
        ws.cell(i+2, 1, value=dummy_data["table_1"]["columns"][i])
        ws.cell(i+2, 2, value=dummy_data["table_1"]["PKs"][i])
        ws["F1"].value = "URL"
        ws["F2"].value = "定義書のURL"

    ws_relation = wb.create_sheet(title="relation")
    ws_relation.column_dimensions["A"].width = 30
    ws_relation.column_dimensions["B"].width = 30
    ws_relation.column_dimensions["C"].width = 40

    ws_relation.cell(1, 1, value="From")
    ws_relation.cell(1, 2, value="To")
    ws_relation.cell(1, 3, value="Description")

    ws_relation.cell(2, 1, value="table_1")
    ws_relation.cell(2, 2, value="join1")
    ws_relation.cell(2, 3, value="company_idでleft join")

    ws_relation.cell(3, 1, value="table_2")
    ws_relation.cell(3, 2, value="join1")
    ws_relation.cell(3, 3, value="")

    ws_relation.cell(4, 1, value="join1")
    ws_relation.cell(4, 2, value="table_3")
    ws_relation.cell(4, 3, value="companyごとに利益を集計")

    ws_relation.cell(5, 1, value="table_4")
    ws_relation.cell(5, 2, value="join2")
    ws_relation.cell(5, 3, value="product_idでleft join")

    ws_relation.cell(6, 1, value="table_5")
    ws_relation.cell(6, 2, value="join2")
    ws_relation.cell(6, 3, value="")

    ws_relation.cell(7, 1, value="join2")
    ws_relation.cell(7, 2, value="table_6")
    ws_relation.cell(7, 3, value="productごとに販売量を集計")

    ws_relation.cell(8, 1, value="table_3")
    ws_relation.cell(8, 2, value="join3")
    ws_relation.cell(8, 3, value="company_id,product_idでleft join")

    ws_relation.cell(9, 1, value="table_6")
    ws_relation.cell(9, 2, value="join3")
    ws_relation.cell(9, 3, value="")

    ws_relation.cell(10, 1, value="join3")
    ws_relation.cell(10, 2, value="table_7")
    ws_relation.cell(10, 3, value="companyごと、productごとの利益を集計")

    table2 = Table(displayName="Relation", ref="A1:C6")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table2.tableStyleInfo = style
    ws_relation.add_table(table2)

    wb.save(filename)
    print(f"Generated {filename}")


def make_excel_sheet(ws, title, n_column, columns, PKs, FKs, comments):
    ws.cell(1, 1, value="Column Name")
    ws.cell(1, 2, value="Primary Key")
    ws.cell(1, 3, value="Foreign Key")
    ws.cell(1, 4, value="Comment")
    for i in range(n_column):
        ws.cell(i+2, 1, value=columns[i])
        ws.cell(i+2, 2, value=PKs[i])
        ws.cell(i+2, 3, value=FKs[i])
        ws.cell(i+2, 4, value=comments[i])
    table = Table(displayName=title, ref=f"A1:D{n_column+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    return table

