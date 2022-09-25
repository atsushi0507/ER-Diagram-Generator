import os
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

class ERDiagram:
    def __init__(self, title="Sample Diagram"):
        self.title = title
        self.obj = f"```plantuml\n@startuml\ntitle {self.title}\n"
        #self.relation = f"```plantuml\n@startuml\nskinparam linetype ortho\n"
        self.relation = f"```plantuml\n@startuml\ntop to bottom direction\n"


    def make_entity(self, entity_name, primary_keys=None, foreign_keys=None):
        self.obj += f"entity {entity_name}" + "{\n"
        if primary_keys is not None:
            if type(primary_keys) == str:
                self.obj += f"* {primary_keys} [PK]\n"
            else:
                for key in primary_keys:
                    self.obj += f"+ {key} [PK]\n"
        self.obj += "--\n"
        if foreign_keys is not None:
            if type(foreign_keys) == str:
                self.obj += f"{foreign_keys} [FK]\n"
            else:
                for key in foreign_keys:
                    self.obj += f"{key} [FK]\n"
        self.obj += "}\n"

    
    def make_relation(self, parent, child, type=4):
        """
        type:
            0: 0 or 1
            1: only 1
            2: 0 or more
            3: 1 or more
            4: arrow
        """
        zero_or_one = "|o--"
        only_one = "||--"
        zero_or_more = "}o--"
        one_or_more = "}|--"
        arrow = "--|>"
        if type == 0:
            marker = zero_or_one
        elif type == 1:
            marker = only_one
        elif type == 2:
            marker = zero_or_more
        elif type == 3:
            marker = one_or_more
        elif type == 4:
            marker = arrow
        self.relation += f"{parent} {marker} {child}" + "\n"


    def load_table(self, load_file):
        self.relation += f"!include {load_file}\n\n"


    def make_entities_from_excel(self, filename):
        book = pd.ExcelFile(filename)
        sheets = book.sheet_names
        for sheet_name in sheets:
            if sheet_name == "relation":
                continue
            df = book.parse(sheet_name=sheet_name, index_col=0)

            df_dict = df.to_dict()
            PKs = []
            FKs = []
            vars = []
            comments = []
            for key, values in df_dict.items():
                for k, value in values.items():
                    if key == "Primary Key" and value == 1.0:
                        PKs.append(k)
                    elif key == "Foreign Key" and value == 1.0:
                        FKs.append(k)

            if len(PKs) == 0:
                PKs = None
            if len(FKs) == 0:
                FKs = None
            if len(vars) == 0:
                vars = None

            self.make_entity(sheet_name, PKs, FKs)
        self.output_table(f"{filename.split('.')[-2]}.md")


    def make_relations_from_excel(self, filename, load_file):
        book = pd.ExcelFile(filename)
        if load_file is not None:
            self.load_table(f"../{load_file}")

        df = book.parse(sheet_name="relation")

        parents = df.Parent.to_list()
        children = df.Child.to_list()

        for i in range(len(parents)):
            self.make_relation(parents[i], children[i])

        self.output_relation(f"{filename.split('.')[-2]}.md")


    def make_all_from_excel(self, filename):
        self.make_entities_from_excel(filename)
        load_file_name = filename.split(".")[-2] + ".md"
        self.make_relations_from_excel(filename, f"tables/{load_file_name}")

    
    def output_table(self, out_name):
        out_dir = "tables"
        if not os.path.isdir(out_dir):
            os.makedirs(out_dir)
        with open(f"{out_dir}/{out_name}", "w") as f:
            self.obj += "@enduml"
            f.write(self.obj)


    def output_relation(self, out_name):
        out_dir = "relation"
        if not os.path.isdir(out_dir):
            os.makedirs(out_dir)
        with open(f"{out_dir}/relation-{out_name}", "w") as f:
            self.relation += "@enduml"
            f.write(self.relation)


def make_excel_sheet(ws, title, n_column, columns, PKs, FKs, comments):
    ws.cell(1, 1, value="Column Name")
    ws.cell(1, 2, value="Primary Key")
    ws.cell(1, 3, value="Foreign Key")
    ws.cell(1, 4, value="Comment")
    for i in range(n_column):
        ws.cell(i+2, 1, value=columns[i])
        ws.cell(i+2, 2, value=PKs[i])
        ws.cell(i+2, 3, value=FKs[i])
        ws.cell(i+2, 4, value=comments[i])
    table = Table(displayName=title, ref=f"A1:D{n_column+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    return table

def generate_excelbook(filename="sample_format.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "sample_table"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 35

    columns = [f"column_{i}" for i in range(10)]
    PKs = [1, 0, 1, 0, 0, 1, 0, 0, 0, 0]
    FKs = [0, 1, 0, 1, 0, 0, 0, 0, 0, 0]
    comments = ["", "Table2の外部キー", "", "Table5の外部キー", "", "", "", "", "", ""]

    table = make_excel_sheet(ws, "Table1", len(columns), columns, PKs, FKs, comments)
    ws.add_table(table)

    ws_child1 = wb.create_sheet(title="child1")
    ws_child1.column_dimensions["A"].width = 30
    ws_child1.column_dimensions["B"].width = 10
    ws_child1.column_dimensions["C"].width = 10
    ws_child1.column_dimensions["D"].width = 35
    columns = [f"column_{i}" for i in range(7)]
    PKs = [0, 1, 0, 0, 0, 0, 0]
    FKs = [0, 0, 0, 1, 0, 0, 0]
    comments = ["" for _ in range(7)]
    table_child1 = make_excel_sheet(ws_child1, "Child1", len(columns), columns, PKs, FKs, comments)
    ws_child1.add_table(table_child1)

    ws_child2 = wb.create_sheet(title="child2")
    ws_child2.column_dimensions["A"].width = 30
    ws_child2.column_dimensions["B"].width = 10
    ws_child2.column_dimensions["C"].width = 10
    ws_child2.column_dimensions["D"].width = 35
    columns = [f"column_{i}" for i in range(4)]
    PKs = [1, 1, 0, 0]
    FKs = [0, 0, 0, 1]
    comments = ["" for _ in range(4)]
    table_child2 = make_excel_sheet(ws_child2, "Child2", len(columns), columns, PKs, FKs, comments)
    ws_child2.add_table(table_child2)

    ws_child3 = wb.create_sheet(title="child3")
    ws_child3.column_dimensions["A"].width = 30
    ws_child3.column_dimensions["B"].width = 10
    ws_child3.column_dimensions["C"].width = 10
    ws_child3.column_dimensions["D"].width = 35
    columns = [f"column_{i}" for i in range(6)]
    PKs = [1, 0, 0, 0, 0, 0]
    FKs = [0, 0, 0, 0, 1, 0]
    comments = ["" for _ in range(6)]
    table_child3 = make_excel_sheet(ws_child3, "Child3", len(columns), columns, PKs, FKs, comments)
    ws_child3.add_table(table_child3)

    ws_child4 = wb.create_sheet(title="child4")
    ws_child4.column_dimensions["A"].width = 30
    ws_child4.column_dimensions["B"].width = 10
    ws_child4.column_dimensions["C"].width = 10
    ws_child4.column_dimensions["D"].width = 35
    columns = [f"column_{i}" for i in range(5)]
    PKs = [1, 1, 0, 1, 0]
    FKs = [0, 0, 0, 1, 0]
    comments = ["" for _ in range(5)]
    table_child4 = make_excel_sheet(ws_child4, "Child4", len(columns), columns, PKs, FKs, comments)
    ws_child4.add_table(table_child4)

    ws_child5 = wb.create_sheet(title="child5")
    ws_child5.column_dimensions["A"].width = 30
    ws_child5.column_dimensions["B"].width = 10
    ws_child5.column_dimensions["C"].width = 10
    ws_child5.column_dimensions["D"].width = 35
    columns = [f"column_{i}" for i in range(4)]
    PKs = [1, 1, 0, 0]
    FKs = [0, 0, 0, 1]
    comments = ["" for _ in range(4)]
    table_child5 = make_excel_sheet(ws_child5, "Child5", len(columns), columns, PKs, FKs, comments)
    ws_child5.add_table(table_child5)

    ws_relation = wb.create_sheet(title="relation")
    ws_relation.column_dimensions["A"].width = 30
    ws_relation.column_dimensions["B"].width = 30
    relations = {
        # parent : child
        "sample_table" : ["child1", "child2", "child3"],
        "child1" : ["child4"],
        "child2" : ["child5"]
    }
    ws_relation.cell(1, 1, value="Parent")
    ws_relation.cell(1, 2, value="Child")
    i = 2
    for key, val in relations.items():
        for v in val:
            ws_relation.cell(i, 1, value=key)
            ws_relation.cell(i, 2, value=v)
            i += 1

    table2 = Table(displayName="Relation", ref="A1:B6")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table2.tableStyleInfo = style

    ws_relation.add_table(table2)

    wb.save(filename)
    print(f"Generated {filename}")


def build_parser():
    description="""
    ER図作成ツールです。実行のヒント:
    python ER_diagram_generator.py --generate_excel (--excel_file [path/to/file])
    python ER_diagram_generator.py --make_tables (--excel_file [path/to/file])
    python ER_diagram_generator.py --make_relations (--excel_file [path/to/file])
    python ER_diagram_generator.py --make_all (--excel_file [path/to/file])
    """
    parser = argparse.ArgumentParser(formatter_class=argparse.RawDescriptionHelpFormatter,
                                     description=description)
    parser.add_argument("-g", "--generate_excel", action="store_true", 
                        help="テーブル定義用のエクセルファイルを生成する。")
    parser.add_argument("-t", "--make_tables", action="store_true", 
                        help="テーブルの一覧を生成する。")
    parser.add_argument("-r", "--make_relations", action="store_true",
                        help="テーブルの関係を作成する。変数は表示しない。")
    parser.add_argument("-a", "--make_all", action="store_true", 
                        help="テーブルの関係を、PKとFK付きで表示する。")
    parser.add_argument("-f", "--excel_file", type=str, default="sample_format.xlsx",
                        help="入力ファイル名を指定する。デフォルトは'sample_format.xlsx'")
    return parser

def main():
    args = build_parser().parse_args()
    er = ERDiagram()
    if args.generate_excel:
        generate_excelbook(args.excel_file)
    if args.make_tables:
        er.make_entities_from_excel(args.excel_file)
    if args.make_relations:
        er.make_relations_from_excel(args.excel_file, None)
    if args.make_all:
        er.make_all_from_excel(args.excel_file)

if __name__ == "__main__":
    main()